import io
from datetime import datetime, timedelta
from flask import Flask, render_template_string, request, redirect, url_for, make_response, session
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'steely_rmi_secure_secret_key_2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///steely_rmi_garage_v8.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
db = SQLAlchemy(app)

class WorkOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    serial_number = db.Column(db.String(50), nullable=False)
    work_order_no = db.Column(db.String(50), nullable=False)
    vehicle_plate = db.Column(db.String(50), nullable=False)
    vehicle_model = db.Column(db.String(100), nullable=False)
    current_reading = db.Column(db.String(50), nullable=False)
    reading_unit = db.Column(db.String(20), nullable=False)
    job_status = db.Column(db.String(50), nullable=False)
    driver_name = db.Column(db.String(100), nullable=False)
    assigned_technicians = db.Column(db.String(200), nullable=False)
    start_datetime = db.Column(db.String(50), nullable=False)
    end_datetime = db.Column(db.String(50), nullable=False)
    maintenance_type = db.Column(db.String(50), nullable=False) 
    work_category = db.Column(db.String(100), nullable=False) 
    description = db.Column(db.Text, nullable=True)
    spare_parts_qty = db.Column(db.Integer, nullable=False, default=0)
    spare_parts_cost = db.Column(db.Float, nullable=False, default=0.0)
    lubricants_volume = db.Column(db.Float, nullable=False, default=0.0)
    lubricants_cost = db.Column(db.Float, nullable=False, default=0.0)
    batteries_cost = db.Column(db.Float, nullable=False, default=0.0)
    tires_cost = db.Column(db.Float, nullable=False, default=0.0)
    effective_work_hours = db.Column(db.Float, nullable=False, default=0.0)
    total_expenditure = db.Column(db.Float, nullable=False, default=0.0)

class SpareInventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    part_name = db.Column(db.String(100), nullable=False)
    spec = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    location = db.Column(db.String(100), nullable=False)

with app.app_context():
    db.create_all()

LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>SteelY R.M.I - Login</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-dark d-flex align-items-center justify-content-center vh-100">
    <div class="card p-4 shadow-lg" style="width: 380px; background: #1e2124; color: white; border-radius: 12px;">
        <h3 class="text-center mb-4 text-primary fw-bold">SteelY R.M.I</h3>
        {% if error %}
            <div class="alert alert-danger py-2 text-center">{{ error }}</div>
        {% endif %}
        <form method="POST" action="/login">
            <div class="mb-3">
                <label class="form-label">Username</label>
                <input type="text" class="form-control" name="username" required autocomplete="off" value="admin">
            </div>
            <div class="mb-3">
                <label class="form-label">Password</label>
                <input type="password" class="form-control" name="password" required value="steely2026">
            </div>
            <button type="submit" class="btn btn-primary w-100 py-2 fw-bold">Login to Dashboard</button>
        </form>
    </div>
</body>
</html>
"""

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>SteelY R.M.I Garage Maintnace dash Bord</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light p-3">
    <div class="container-fluid px-3">
        <div class="card shadow-sm p-3 mb-3 bg-white">
            <div class="row align-items-center g-2">
                <div class="col-md-4">
                    <h4 class="text-primary fw-bold mb-0">SteelY R.M.I Garage Maintenance</h4>
                    <small class="text-muted">Integrated Work Time, Consumables & Maintenance Platform</small>
                </div>
                <div class="col-md-8 d-flex justify-content-md-end align-items-center gap-2 flex-wrap">
                    <div class="text-start text-md-end">
                        <span class="badge bg-secondary">Dinberu Tefera</span><br>
                        <small class="text-muted fw-bold" style="font-size: 9px;">HEAD OF MECHANICAL WORKSHOP AND GARAGE</small>
                    </div>
                    <a href="/export/master_report" class="btn btn-success btn-sm fw-bold">📊 All-in-one-master report export to excel</a>
                    <a href="/logout" class="btn btn-danger btn-sm fw-bold">🚪 Logout</a>
                </div>
            </div>
        </div>
        
        <div class="mb-3 d-flex gap-2">
            <button class="btn btn-primary btn-sm fw-bold" data-bs-toggle="modal" data-bs-target="#addWorkOrderModal">+ Create New Work Order</button>
            <button class="btn btn-dark btn-sm fw-bold" data-bs-toggle="modal" data-bs-target="#addSpareModal">+ Store Spare Inventory</button>
        </div>

        <div class="row mb-3">
            <div class="col-md-6 mb-2">
                <div class="card shadow-sm h-100">
                    <div class="card-header bg-secondary text-white py-2">
                        <h5 class="mb-0 fs-6">WEEKLY SUMMARY (LAST 7 DAYS)</h5>
                    </div>
                    <div class="card-body">
                        <p class="fw-bold mb-2">Total Jobs Executed: <span class="text-primary">{{ weekly_jobs }}</span></p>
                        <ul class="list-unstyled ms-3 mb-2 small">
                            <li>• Preventive Maintenance (PM): <strong>{{ weekly_pm }}</strong></li>
                            <li>• Corrective Maintenance (CM): <strong>{{ weekly_cm }}</strong></li>
                            <li>• Inspection & Checkup: <strong>{{ weekly_insp }}</strong></li>
                        </ul>
                        <p class="fw-bold mb-2 text-primary">Total Effective Work Time: {{ "%.1f"|format(weekly_hours) }} hrs</p>
                        <hr class="my-2">
                        <div class="row small text-muted">
                            <div class="col-6">Spare Qty: <strong>{{ weekly_spare_qty }} Pcs</strong></div>
                            <div class="col-6">Spare Cost: <strong>ETB {{ "%.2f"|format(weekly_spare_cost) }}</strong></div>
                            <div class="col-6">Lubricants: <strong>{{ "%.1f"|format(weekly_lube_vol) }} L</strong></div>
                            <div class="col-6">Lube Cost: <strong>ETB {{ "%.2f"|format(weekly_lube_cost) }}</strong></div>
                            <div class="col-6">Batteries Cost: <strong>ETB {{ "%.2f"|format(weekly_batt_cost) }}</strong></div>
                            <div class="col-6">Tires Cost: <strong>ETB {{ "%.2f"|format(weekly_tire_cost) }}</strong></div>
                        </div>
                        <div class="mt-2 pt-2 border-top fw-bold text-dark">
                            Total Expenditure: ETB {{ "%.2f"|format(weekly_total_exp) }}
                        </div>
                    </div>
                </div>
            </div>

            <div class="col-md-6 mb-2">
                <div class="card shadow-sm h-100">
                    <div class="card-header bg-primary text-white py-2">
                        <h5 class="mb-0 fs-6">MONTHLY SUMMARY (LAST 30 DAYS)</h5>
                    </div>
                    <div class="card-body">
                        <p class="fw-bold mb-2">Total Jobs Executed: <span class="text-primary">{{ monthly_jobs }}</span></p>
                        <ul class="list-unstyled ms-3 mb-2 small">
                            <li>• Preventive Maintenance (PM): <strong>{{ monthly_pm }}</strong></li>
                            <li>• Corrective Maintenance (CM): <strong>{{ monthly_cm }}</strong></li>
                            <li>• Inspection & Checkup: <strong>{{ monthly_insp }}</strong></li>
                        </ul>
                        <p class="fw-bold mb-2 text-primary">Total Effective Work Time: {{ "%.1f"|format(monthly_hours) }} hrs</p>
                        <hr class="my-2">
                        <div class="row small text-muted">
                            <div class="col-6">Spare Qty: <strong>{{ monthly_spare_qty }} Pcs</strong></div>
                            <div class="col-6">Spare Cost: <strong>ETB {{ "%.2f"|format(monthly_spare_cost) }}</strong></div>
                            <div class="col-6">Lubricants: <strong>{{ "%.1f"|format(monthly_lube_vol) }} L</strong></div>
                            <div class="col-6">Lube Cost: <strong>ETB {{ "%.2f"|format(monthly_lube_cost) }}</strong></div>
                            <div class="col-6">Batteries Cost: <strong>ETB {{ "%.2f"|format(monthly_batt_cost) }}</strong></div>
                            <div class="col-6">Tires Cost: <strong>ETB {{ "%.2f"|format(monthly_tire_cost) }}</strong></div>
                        </div>
                        <div class="mt-2 pt-2 border-top fw-bold text-success">
                            Total Expenditure: ETB {{ "%.2f"|format(monthly_total_exp) }}
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <div class="card shadow-sm mb-3">
            <div class="card-header bg-secondary text-white py-2">
                <h5 class="mb-0 fs-6">Store Spare Inventory</h5>
            </div>
            <div class="card-body p-2">
                <table class="table table-bordered table-hover align-middle mb-0 small">
                    <thead class="table-light">
                        <tr>
                            <th>ID</th>
                            <th>Part Name</th>
                            <th>Specification (Spec)</th>
                            <th>Available Quantity</th>
                            <th>Location</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for item in inventory_items %}
                        <tr>
                            <td>{{ item.id }}</td>
                            <td>{{ item.part_name }}</td>
                            <td>{{ item.spec }}</td>
                            <td class="fw-bold {% if item.quantity < 5 %}text-danger{% else %}text-success{% endif %}">{{ item.quantity }}</td>
                            <td>{{ item.location }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>

        <div class="card shadow-sm">
            <div class="card-header bg-primary text-white py-2">
                <h5 class="mb-0 fs-6">Maintenance Execution & Work Time Log</h5>
            </div>
            <div class="card-body p-2">
                <table class="table table-bordered table-hover align-middle mb-0 small">
                    <thead class="table-secondary">
                        <tr>
                            <th>ID / S/N</th>
                            <th>Work Order No</th>
                            <th>Vehicle Model & Plate</th>
                            <th>Job Status</th>
                            <th>Maintenance Type</th>
                            <th>Work Category & Description</th>
                            <th>Assigned Technicians</th>
                            <th>Start / End Time</th>
                            <th>Total Cost (ETB)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for wo in work_orders %}
                        <tr>
                            <td>{{ wo.serial_number }}</td>
                            <td>{{ wo.work_order_no }}</td>
                            <td>{{ wo.vehicle_model }} ({{ wo.vehicle_plate }})</td>
                            <td>
                                {% if wo.job_status == 'Completed' %}
                                    <span class="badge bg-success">Completed</span>
                                {% else %}
                                    <span class="badge bg-warning text-dark">{{ wo.job_status }}</span>
                                {% endif %}
                            </td>
                            <td>
                                <span class="badge bg-info text-dark">{{ wo.maintenance_type }}</span>
                            </td>
                            <td>{{ wo.work_category }}</td>
                            <td>{{ wo.assigned_technicians }}</td>
                            <td><small>{{ wo.start_datetime }} to {{ wo.end_datetime }}</small></td>
                            <td class="fw-bold text-success">{{ "%.2f"|format(wo.total_expenditure) }} ETB</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <div class="modal fade" id="addWorkOrderModal" tabindex="-1">
        <div class="modal-dialog modal-lg">
            <div class="modal-content">
                <form method="POST" action="/add_work_order">
                    <div class="modal-header bg-primary text-white">
                        <h5 class="modal-title">Create New Work Order</h5>
                        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
                    </div>
                    <div class="modal-body">
                        <div class="row g-3">
                            <div class="col-md-4">
                                <label class="form-label">Serial Number (S/N)</label>
                                <input type="text" class="form-control" name="serial_number" required value="SN-001">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Work Order No</label>
                                <input type="text" class="form-control" name="work_order_no" required value="WO-2026-01">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Vehicle Plate Number</label>
                                <input type="text" class="form-control" name="vehicle_plate" required placeholder="e.g. AA-3-12345">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Vehicle Type / Model</label>
                                <input type="text" class="form-control" name="vehicle_model" required placeholder="e.g. Sino Truck 371">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Current Reading</label>
                                <input type="text" class="form-control" name="current_reading" required placeholder="e.g. 125000">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Reading Unit</label>
                                <input type="text" class="form-control" name="reading_unit" required value="KM">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Job Status</label>
                                <select class="form-select" name="job_status">
                                    <option value="Completed">Completed</option>
                                    <option value="In Progress">In Progress</option>
                                    <option value="Pending">Pending</option>
                                </select>
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Driver Name</label>
                                <input type="text" class="form-control" name="driver_name" required placeholder="Driver Name">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Assigned Technicians / Mechanics</label>
                                <input type="text" class="form-control" name="assigned_technicians" required value="Ato Mihret, Dinberu Tefera">
                            </div>
                            <div class="col-md-6">
                                <label class="form-label">Start Date & Time</label>
                                <input type="datetime-local" class="form-control" name="start_datetime" required>
                            </div>
                            <div class="col-md-6">
                                <label class="form-label">End Date & Time</label>
                                <input type="datetime-local" class="form-control" name="end_datetime" required>
                            </div>
                            <div class="col-md-4">
                                <label class="form-label fw-bold text-danger">Maintenance Type</label>
                                <select class="form-select border-danger fw-bold text-primary" name="maintenance_type" required>
                                    <option value="CM">CM (Corrective Maintenance)</option>
                                    <option value="PM">PM (Preventive Maintenance)</option>
                                    <option value="Inspection & Check">Inspection & Check</option>
                                </select>
                            </div>
                            <div class="col-md-8">
                                <label class="form-label fw-bold">Work Category & Description</label>
                                <input type="text" class="form-control mb-1" name="work_category" required placeholder="e.g. Engine Maintenance">
                            </div>
                            <div class="col-12">
                                <textarea class="form-control" name="description" rows="2" placeholder="Detailed work description or notes..."></textarea>
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Spare Parts Qty (Pcs)</label>
                                <input type="number" class="form-control" name="spare_parts_qty" value="0" min="0">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Spare Parts Cost (ETB)</label>
                                <input type="number" step="0.01" class="form-control" name="spare_parts_cost" value="0.00" min="0">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Lubricants Volume (Liters)</label>
                                <input type="number" step="0.1" class="form-control" name="lubricants_volume" value="0.0" min="0">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Lubricants Cost (ETB)</label>
                                <input type="number" step="0.01" class="form-control" name="lubricants_cost" value="0.00" min="0">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Batteries Cost (ETB)</label>
                                <input type="number" step="0.01" class="form-control" name="batteries_cost" value="0.00" min="0">
                            </div>
                            <div class="col-md-4">
                                <label class="form-label">Tires Cost (ETB)</label>
                                <input type="number" step="0.01" class="form-control" name="tires_cost" value="0.00" min="0">
                            </div>
                        </div>
                    </div>
                    <div class="modal-footer">
                        <button type="submit" class="btn btn-primary">Save Work Order</button>
                    </div>
                </form>
            </div>
        </div>
    </div>

    <div class="modal fade" id="addSpareModal" tabindex="-1">
        <div class="modal-dialog">
            <div class="modal-content">
                <form method="POST" action="/add_spare">
                    <div class="modal-header bg-dark text-white">
                        <h5 class="modal-title">Add Store Spare Inventory</h5>
                        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
                    </div>
                    <div class="modal-body">
                        <div class="mb-3">
                            <label class="form-label">Part Name</label>
                            <input type="text" class="form-control" name="part_name" required placeholder="Part Name">
                        </div>
                        <div class="mb-3">
                            <label class="form-label">Specification (Spec)</label>
                            <input type="text" class="form-control" name="spec" required placeholder="Specification">
                        </div>
                        <div class="mb-3">
                            <label class="form-label">Quantity</label>
                            <input type="number" class="form-control" name="quantity" required min="1" value="10">
                        </div>
                        <div class="mb-3">
                            <label class="form-label">Location</label>
                            <input type="text" class="form-control" name="location" required value="Main Store">
                        </div>
                    </div>
                    <div class="modal-footer">
                        <button type="submit" class="btn btn-dark">Save Spare</button>
                    </div>
                </form>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

@app.route('/')
def index():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'steely2026':
            session.clear()
            session['logged_in'] = True
            return redirect(url_for('dashboard'))
        else:
            error = 'Invalid username or password!'
    return render_template_string(LOGIN_HTML, error=error)

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    work_orders = WorkOrder.query.all()
    inventory_items = SpareInventory.query.all()

    now = datetime.now()
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    weekly_jobs = weekly_pm = weekly_cm = weekly_insp = 0
    weekly_hours = weekly_spare_qty = weekly_spare_cost = 0.0
    weekly_lube_vol = weekly_lube_cost = weekly_batt_cost = weekly_tire_cost = weekly_total_exp = 0.0

    monthly_jobs = monthly_pm = monthly_cm = monthly_insp = 0
    monthly_hours = monthly_spare_qty = monthly_spare_cost = 0.0
    monthly_lube_vol = monthly_lube_cost = monthly_batt_cost = monthly_tire_cost = monthly_total_exp = 0.0

    for wo in work_orders:
        try:
            wo_date = datetime.strptime(wo.start_datetime, '%Y-%m-%dT%H:%M')
        except:
            try:
                wo_date = datetime.strptime(wo.start_datetime, '%Y-%m-%d %H:%M')
            except:
                wo_date = now

        if wo_date >= month_ago:
            monthly_jobs += 1
            if wo.maintenance_type == 'PM': monthly_pm += 1
            elif wo.maintenance_type == 'CM': monthly_cm += 1
            else: monthly_insp += 1
            monthly_hours += wo.effective_work_hours
            monthly_spare_qty += wo.spare_parts_qty
            monthly_spare_cost += wo.spare_parts_cost
            monthly_lube_vol += wo.lubricants_volume
            monthly_lube_cost += wo.lubricants_cost
            monthly_batt_cost += wo.batteries_cost
            monthly_tire_cost += wo.tires_cost
            monthly_total_exp += wo.total_expenditure

        if wo_date >= week_ago:
            weekly_jobs += 1
            if wo.maintenance_type == 'PM': weekly_pm += 1
            elif wo.maintenance_type == 'CM': weekly_cm += 1
            else: weekly_insp += 1
            weekly_hours += wo.effective_work_hours
            weekly_spare_qty += wo.spare_parts_qty
            weekly_spare_cost += wo.spare_parts_cost
            weekly_lube_vol += wo.lubricants_volume
            weekly_lube_cost += wo.lubricants_cost
            weekly_batt_cost += wo.batteries_cost
            weekly_tire_cost += wo.tires_cost
            weekly_total_exp += wo.total_expenditure

    response = make_response(render_template_string(
        DASHBOARD_HTML, 
        work_orders=work_orders, 
        inventory_items=inventory_items,
        weekly_jobs=weekly_jobs, weekly_pm=weekly_pm, weekly_cm=weekly_cm, weekly_insp=weekly_insp,
        weekly_hours=weekly_hours, weekly_spare_qty=weekly_spare_qty, weekly_spare_cost=weekly_spare_cost,
        weekly_lube_vol=weekly_lube_vol, weekly_lube_cost=weekly_lube_cost, weekly_batt_cost=weekly_batt_cost,
        weekly_tire_cost=weekly_tire_cost, weekly_total_exp=weekly_total_exp,
        monthly_jobs=monthly_jobs, monthly_pm=monthly_pm, monthly_cm=monthly_cm, monthly_insp=monthly_insp,
        monthly_hours=monthly_hours, monthly_spare_qty=monthly_spare_qty, monthly_spare_cost=monthly_spare_cost,
        monthly_lube_vol=monthly_lube_vol, monthly_lube_cost=monthly_lube_cost, monthly_batt_cost=monthly_batt_cost,
        monthly_tire_cost=monthly_tire_cost, monthly_total_exp=monthly_total_exp
    ))
    
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route('/logout')
def logout():
    session.clear()
    response = make_response(redirect(url_for('login')))
    response.delete_cookie('session')
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response

@app.route('/add_spare', methods=['POST'])
def add_spare():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    part_name = request.form.get('part_name')
    spec = request.form.get('spec')
    quantity = int(request.form.get('quantity'))
    location = request.form.get('location')
    
    new_spare = SpareInventory(part_name=part_name, spec=spec, quantity=quantity, location=location)
    db.session.add(new_spare)
    db.session.commit()
    return redirect(url_for('dashboard'))

@app.route('/add_work_order', methods=['POST'])
def add_work_order():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    serial_number = request.form.get('serial_number')
    work_order_no = request.form.get('work_order_no')
    vehicle_plate = request.form.get('vehicle_plate')
    vehicle_model = request.form.get('vehicle_model')
    current_reading = request.form.get('current_reading')
    reading_unit = request.form.get('reading_unit')
    job_status = request.form.get('job_status')
    driver_name = request.form.get('driver_name')
    assigned_technicians = request.form.get('assigned_technicians')
    start_datetime = request.form.get('start_datetime')
    end_datetime = request.form.get('end_datetime')
    maintenance_type = request.form.get('maintenance_type')
    work_category = request.form.get('work_category')
    description = request.form.get('description')
    
    spare_parts_qty = int(request.form.get('spare_parts_qty', 0))
    spare_parts_cost = float(request.form.get('spare_parts_cost', 0.0))
    lubricants_volume = float(request.form.get('lubricants_volume', 0.0))
    lubricants_cost = float(request.form.get('lubricants_cost', 0.0))
    batteries_cost = float(request.form.get('batteries_cost', 0.0))
    tires_cost = float(request.form.get('tires_cost', 0.0))
    
    effective_hours = 0.0
    try:
        s_dt = datetime.strptime(start_datetime, '%Y-%m-%dT%H:%M')
        e_dt = datetime.strptime(end_datetime, '%Y-%m-%dT%H:%M')
        diff = (e_dt - s_dt).total_seconds() / 3600.0
        effective_hours = max(0.0, diff)
    except:
        effective_hours = 2.0

    total_expenditure = spare_parts_cost + lubricants_cost + batteries_cost + tires_cost
    
    new_wo = WorkOrder(
        serial_number=serial_number,
        work_order_no=work_order_no,
        vehicle_plate=vehicle_plate,
        vehicle_model=vehicle_model,
        current_reading=current_reading,
        reading_unit=reading_unit,
        job_status=job_status,
        driver_name=driver_name,
        assigned_technicians=assigned_technicians,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        maintenance_type=maintenance_type,
        work_category=work_category,
        description=description,
        spare_parts_qty=spare_parts_qty,
        spare_parts_cost=spare_parts_cost,
        lubricants_volume=lubricants_volume,
        lubricants_cost=lubricants_cost,
        batteries_cost=batteries_cost,
        tires_cost=tires_cost,
        effective_work_hours=effective_hours,
        total_expenditure=total_expenditure
    )
    db.session.add(new_wo)
    db.session.commit()
    return redirect(url_for('dashboard'))

@app.route('/export/master_report')
def export_master_report():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    try:
        work_orders = WorkOrder.query.all()
        inventory_items = SpareInventory.query.all()
        
        wb = openpyxl.Workbook()
        
        # Setup sheets
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_wo = wb.create_sheet(title="Work Orders & Maintenance Logs")
        ws_inv = wb.create_sheet(title="Spare Inventory")

        for ws in [ws_summary, ws_wo, ws_inv]:
            ws.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
        font_normal = Font(name="Calibri", size=11, color="000000")
        
        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_sub = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        border_thin = Side(border_style="thin", color="CBD5E1")
        border_thick = Side(border_style="medium", color="1E3A8A")
        border_double = Side(border_style="double", color="1E3A8A")
        
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thick)
        total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

        # --- TAB 1: EXECUTIVE SUMMARY ---
        ws_summary.merge_cells("A1:D2")
        t_cell = ws_summary["A1"]
        t_cell.value = "STEELY R.M.I GARAGE & WORKSHOP MAINTENANCE MASTER SUMMARY"
        t_cell.font = font_title
        t_cell.fill = fill_header
        t_cell.alignment = align_center
        
        ws_summary["A4"] = "Prepared By:"
        ws_summary["B4"] = "Dinberu Tefera (Head of Mechanical Workshop and Garage)"
        ws_summary["A4"].font = font_bold
        ws_summary["B4"].font = font_normal

        headers_sum = ["Metric Category", "All-Time Cumulative"]
        for col_idx, h_text in enumerate(headers_sum, start=1):
            c = ws_summary.cell(row=6, column=col_idx, value=h_text)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_center
            c.border = header_border

        max_wo_row = max(len(work_orders) + 2, 3)
        summary_rows = [
            ("Total Jobs Executed", f"=COUNTA('Work Orders & Maintenance Logs'!B3:B{max_wo_row})"),
            ("Preventive Maintenance (PM)", f"=COUNTIF('Work Orders & Maintenance Logs'!L3:L{max_wo_row}, \"PM\")"),
            ("Corrective Maintenance (CM)", f"=COUNTIF('Work Orders & Maintenance Logs'!L3:L{max_wo_row}, \"CM\")"),
            ("Inspection & Checkup", f"=COUNTIF('Work Orders & Maintenance Logs'!L3:L{max_wo_row}, \"Inspection & Check\")"),
            ("Total Effective Work Hours (hrs)", f"=SUM('Work Orders & Maintenance Logs'!U3:U{max_wo_row})"),
            ("Spare Parts Cost (ETB)", f"=SUM('Work Orders & Maintenance Logs'!P3:P{max_wo_row})"),
            ("Lubricants Cost (ETB)", f"=SUM('Work Orders & Maintenance Logs'!R3:R{max_wo_row})"),
            ("Batteries Cost (ETB)", f"=SUM('Work Orders & Maintenance Logs'!S3:S{max_wo_row})"),
            ("Tires Cost (ETB)", f"=SUM('Work Orders & Maintenance Logs'!T3:T{max_wo_row})"),
            ("Total Expenditure (ETB)", f"=SUM('Work Orders & Maintenance Logs'!V3:V{max_wo_row})"),
        ]

        for idx, (m_name, m_formula) in enumerate(summary_rows, start=7):
            c1 = ws_summary.cell(row=idx, column=1, value=m_name)
            c1.font = font_bold
            c1.border = cell_border
            c1.alignment = align_left
            
            c2 = ws_summary.cell(row=idx, column=2, value=m_formula)
            c2.font = font_normal
            c2.border = cell_border
            c2.alignment = align_right
            if "Cost" in m_name or "Expenditure" in m_name:
                c2.number_format = '#,##0.00'
            elif "Hours" in m_name:
                c2.number_format = '#,##0.0'
            else:
                c2.number_format = '#,##0'

        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 25

        # --- TAB 2: WORK ORDERS & MAINTENANCE LOGS ---
        wo_headers = [
            "Serial Number", "Work Order No", "Vehicle Plate", "Vehicle Model", 
            "Current Reading", "Reading Unit", "Job Status", "Driver Name", 
            "Assigned Technicians", "Start Time", "End Time", 
            "Maintenance Type", "Work Category", "Description", 
            "Spare Qty", "Spare Cost (ETB)", "Lube Vol (L)", "Lube Cost (ETB)", 
            "Batt Cost (ETB)", "Tire Cost (ETB)", "Effective Hours", "Total Expenditure (ETB)"
        ]

        ws_wo.merge_cells("A1:V1")
        t_wo = ws_wo["A1"]
        t_wo.value = "SECTION 1: WORK ORDERS & MAINTENANCE EXECUTION LOGS"
        t_wo.font = font_title
        t_wo.fill = fill_header
        t_wo.alignment = align_center

        for col_idx, text in enumerate(wo_headers, start=1):
            c = ws_wo.cell(row=2, column=col_idx, value=text)
            c.font = font_header
            c.fill = fill_sub
            c.alignment = align_center
            c.border = header_border

        for row_idx, wo in enumerate(work_orders, start=3):
            tot_form = f"=P{row_idx}+R{row_idx}+S{row_idx}+T{row_idx}"
            row_data = [
                wo.serial_number, wo.work_order_no, wo.vehicle_plate, wo.vehicle_model, 
                wo.current_reading, wo.reading_unit, wo.job_status, wo.driver_name, 
                wo.assigned_technicians, wo.start_datetime, wo.end_datetime, 
                wo.maintenance_type, wo.work_category, wo.description, 
                wo.spare_parts_qty, wo.spare_parts_cost, wo.lubricants_volume, wo.lubricants_cost, 
                wo.batteries_cost, wo.tires_cost, wo.effective_work_hours, tot_form
            ]
            for col_idx, val in enumerate(row_data, start=1):
                c = ws_wo.cell(row=row_idx, column=col_idx, value=val)
                c.font = font_normal
                c.border = cell_border
                if col_idx in [15, 17]:
                    c.alignment = align_right
                    c.number_format = '#,##0.0' if col_idx == 17 else '#,##0'
                elif col_idx in [16, 18, 19, 20, 22]:
                    c.alignment = align_right
                    c.number_format = '#,##0.00'
                elif col_idx == 21:
                    c.alignment = align_right
                    c.number_format = '#,##0.0'
                else:
                    c.alignment = align_left
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

        # Totals row for Work Orders
        tot_r_wo = len(work_orders) + 3
        ws_wo.cell(row=tot_r_wo, column=1, value="TOTAL").font = font_bold
        ws_wo.cell(row=tot_r_wo, column=1).border = total_border
        for c_idx in range(2, 23):
            c = ws_wo.cell(row=tot_r_wo, column=c_idx)
            c.border = total_border
            col_ltr = get_column_letter(c_idx)
            if c_idx in [15, 16, 17, 18, 19, 20, 21, 22]:
                c.value = f"=SUM({col_ltr}3:{col_ltr}{tot_r_wo-1})"
                c.font = font_bold
                c.alignment = align_right
                c.number_format = '#,##0.00' if c_idx not in [15, 17, 21] else '#,##0.0'

        for col in ws_wo.columns:
            max_l = max(len(str(cell.value or '')) for cell in col)
            ws_wo.column_dimensions[get_column_letter(col[0].column)].width = max(max_l + 3, 15)

        # --- TAB 3: SPARE INVENTORY ---
        inv_headers = ["ID", "Part Name", "Specification (Spec)", "Available Quantity", "Location"]
        ws_inv.merge_cells("A1:E1")
        t_inv = ws_inv["A1"]
        t_inv.value = "SECTION 2: STORE SPARE INVENTORY MASTER"
        t_inv.font = font_title
        t_inv.fill = fill_header
        t_inv.alignment = align_center

        for col_idx, text in enumerate(inv_headers, start=1):
            c = ws_inv.cell(row=2, column=col_idx, value=text)
            c.font = font_header
            c.fill = fill_sub
            c.alignment = align_center
            c.border = header_border

        for row_idx, item in enumerate(inventory_items, start=3):
            item_data = [item.id, item.part_name, item.spec, item.quantity, item.location]
            for col_idx, val in enumerate(item_data, start=1):
                c = ws_inv.cell(row=row_idx, column=col_idx, value=val)
                c.font = font_normal
                c.border = cell_border
                if col_idx == 4:
                    c.alignment = align_right
                    c.number_format = '#,##0'
                else:
                    c.alignment = align_left
                if row_idx % 2 == 0:
                    c.fill = fill_zebra

        for col in ws_inv.columns:
            max_l = max(len(str(cell.value or '')) for cell in col)
            ws_inv.column_dimensions[get_column_letter(col[0].column)].width = max(max_l + 5, 20)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = make_response(output.read())
        resp.headers["Content-Disposition"] = "attachment; filename=steely_rmi_master_report_2026.xlsx"
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return resp
    except Exception as e:
        return f"Export Error: {str(e)}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
